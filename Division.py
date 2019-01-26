# -*- coding: utf-8 -*-

from openpyxl import load_workbook
import csv
import datetime
import os.path
import glob
import Library
from shutil import copyfile

def MainDivision(rawdata_path, rawdata_date):
    try:
        print('************ BEGIN ************')
        print()
        print('Step 1: Processing Template and Raw Data')
        try:
            
            #1.import template
            wb_file_name = 'Report\AG_Performance_Template.xlsx'
            wb_file = Library.getXlsxFile(wb_file_name, [])
            print('  -Loading file: %s ' % wb_file, end="")
            wb = load_workbook(filename = wb_file)
            wb_sheet = wb["Performance"]
            #ReportDate = wb_sheet.cell(column=1, row=1).value
            ReportDate = rawdata_date if rawdata_date!=None else wb_sheet.cell(column=1, row=1).value
            ReportDateSimpleArr = str(ReportDate).split(' ',1)[0].split('-',2)
            ReportDateSimple = ReportDateSimpleArr[1]+ReportDateSimpleArr[2]
            ReportDateStr = ReportDateSimpleArr[0]+"/"+ReportDateSimpleArr[1]+"/"+ReportDateSimpleArr[2]
            ReportDateStrNoZero = ReportDateSimpleArr[0]+"/"+str(int(ReportDateSimpleArr[1]))+"/"+str(int(ReportDateSimpleArr[2]))
            
            for checksheet in wb.sheetnames:
                if checksheet!="分時表":
                    wb.remove(wb[checksheet])
            wb_sheet = wb["分時表"]
            print(' => Completed')

            #2.import data source
            #Skill1
            print('  -Loading file:', end="")
            try:
                table_skill1_file_name='間隔skill總結-SKILL1'                
                table_skill1_file, table_skill1_list=Library.getCsvFile2(rawdata_path + '\**\*%s*.xls' % table_skill1_file_name if rawdata_path!=None else None, 'Report\RAWDATA\**\*%s*.xls' % table_skill1_file_name, [[1,1, ReportDateStrNoZero], [1,2, 'OPPO SKILL 1']])
                if table_skill1_list!=None: print(' %s => Completed' % table_skill1_file)
                else: print(' Finding "%s\" => Failed' % table_skill1_file_name)
            except Exception as e:
                pass
            
            #Skill2
            print('  -Loading file:', end="")
            try:
                table_skill2_file_name='間隔skill總結-SKILL2'                
                table_skill2_file, table_skill2_list=Library.getCsvFile2(rawdata_path + '\**\*%s*.xls' % table_skill2_file_name if rawdata_path!=None else None, 'Report\RAWDATA\**\*%s*.xls' % table_skill2_file_name, [[1,1, ReportDateStrNoZero], [1,2, 'OPPO SKILL 2']])
                if table_skill2_list!=None: print(' %s => Completed' % table_skill2_file)
                else: print(' Finding "%s\" => Failed' % table_skill2_file_name)
            except Exception as e:
                pass

            #Skill3
            print('  -Loading file:', end="")
            try:
                table_skill3_file_name='間隔skill總結-SKILL3'                
                table_skill3_file, table_skill3_list=Library.getCsvFile2(rawdata_path + '\**\*%s*.xls' % table_skill3_file_name if rawdata_path!=None else None, 'Report\RAWDATA\**\*%s*.xls' % table_skill3_file_name, [[1,1, ReportDateStrNoZero], [1,2, 'OPPO SKILL 3']])
                if table_skill3_list!=None: print(' %s => Completed' % table_skill3_file)
                else: print(' Finding "%s\" => Failed' % table_skill3_file_name)
            except Exception as e:
                pass

            #Skill4
            print('  -Loading file:', end="")
            try:
                table_skill4_file_name='間隔skill總結-SKILL4'                
                table_skill4_file, table_skill4_list=Library.getCsvFile2(rawdata_path + '\**\*%s*.xls' % table_skill4_file_name if rawdata_path!=None else None, 'Report\RAWDATA\**\*%s*.xls' % table_skill4_file_name, [[1,1, ReportDateStrNoZero], [1,2, 'OPPO SKILL 4']])
                if table_skill4_list!=None: print(' %s => Completed' % table_skill4_file)
                else: print(' Finding "%s\" => Failed' % table_skill4_file_name)
            except Exception as e:
                pass

            #VM
            #print('  -Loading file:', end="")
            #try:
            #    table_VM_file_name='間隔skill總結-VM'                
            #    table_VM_file, table_VM_list=Library.getCsvFile2(rawdata_path + '\**\*%s*.xls' % table_VM_file_name if rawdata_path!=None else None, 'Report\RAWDATA\**\*%s*.xls' % table_VM_file_name, [[1,0, ReportDateStrNoZero], [1,2, 'OPPO to IVR 2040000']])
            #    if table_VM_list!=None: print(' %s => Completed' % table_VM_file)
            #    else: print(' Finding "%s\" => Failed' % table_VM_file_name)
            #except Exception as e:
            #    pass

            #VM Version2 from websystem - 20190112
            print('  -Loading file:', end="")
            try:
                table_VM_file_name=Library.getVoiceMail(ReportDate)
                if table_VM_file_name==None: table_VM_file_name='留電報表'                
                table_VM_file, table_VM_list=Library.getCsvFile2(None, table_VM_file_name, [])
                if table_VM_list!=None: print(' %s => Completed' % table_VM_file)
                else: print(' Finding "%s\" => Failed' % table_VM_file_name)
                #print(table_VM_list)
            except Exception as e:
                #print(e)
                pass

            #for skill1 in table_skill1_list:
            #    print(Library.convertTimeDivisionFull(skill1[0], skill1[1]))


            #3.Processing Report
            print()
            print('Step 2: Processing report calculation')

            wb_sheet.cell(column=1, row=2).value=ReportDateStr
            Totaltalkingsec=0
            TotalACD=0
            TotalACW=0
            TotalHoldsec=0
            TotalHold=0
            TotalAcw=0
            TotalPickupSec=0
            for rows in range(3, wb_sheet.max_row+1):
                FullDataTime = ReportDateStr + ' ' + wb_sheet.cell(column=2, row=rows).value #日期/時間
                print('  -Loading Time: %s' % FullDataTime, end="")
                try:
                    wb_sheet.cell(column=1, row=rows).value=FullDataTime
                    wb_sheet.cell(column=4, row=rows).value=0 #ACD通話
                    wb_sheet.cell(column=5, row=rows).value=0 #掛斷通話
                    wb_sheet.cell(column=6, row=rows).value=0 #Voice Mail
                    wb_sheet.cell(column=10, row=rows).value=0 #平均位置值班
                    talkingsec=0
                    holdsec=0 #等候時間
                    hold=0 #等候通話
                    acwsec=0 #ACW
                    pickupsec=0 #接聽時間
                    #Skill1
                    try:
                        for skill1 in table_skill1_list:
                            if wb_sheet.cell(column=2, row=rows).value==Library.convertTimeDivisionFull(skill1[0], skill1[1]):
                                wb_sheet.cell(column=4, row=rows).value+=int(skill1[4]) #ACD通話
                                wb_sheet.cell(column=5, row=rows).value+=int(skill1[7]) #掛斷通話
                                wb_sheet.cell(column=10, row=rows).value+=int(skill1[17]) #平均位置值班
                                holdsec+=int(skill1[20]) #等候時間
                                hold+=int(skill1[21]) #等候通話
                                TotalACD+=int(skill1[4])
                                talkingsec+=float(skill1[4])*float(skill1[5]) #ACD通話*平均ACD時間
                                acwsec+=float(skill1[4])*float(skill1[6]) #ACD通話*平均ACW時間
                                pickupsec+=float(skill1[4])*float(skill1[2]) #ACD通話*平均速度接聽
                                break
                    except: print(' => Missing Skill1', end="")

                    #Skill2
                    try:
                        for skill2 in table_skill2_list:                
                            if wb_sheet.cell(column=2, row=rows).value==Library.convertTimeDivisionFull(skill2[0], skill2[1]):
                                wb_sheet.cell(column=4, row=rows).value+=int(skill2[4]) #ACD通話
                                wb_sheet.cell(column=5, row=rows).value+=int(skill2[7]) #掛斷通話
                                holdsec+=int(skill2[20]) #等候時間
                                hold+=int(skill2[21]) #等候通話
                                TotalACD+=int(skill2[4])
                                talkingsec+=float(skill2[4])*float(skill2[5]) #ACD通話*平均ACD時間
                                acwsec+=float(skill2[4])*float(skill2[6]) #ACD通話*平均ACW時間
                                pickupsec+=float(skill2[4])*float(skill2[2]) #ACD通話*平均速度接聽
                                break
                    except: print(' => Missing Skill2', end="")

                    #Skill3
                    try:
                        for skill3 in table_skill3_list:
                            if wb_sheet.cell(column=2, row=rows).value==Library.convertTimeDivisionFull(skill3[0], skill3[1]):
                                wb_sheet.cell(column=4, row=rows).value+=int(skill3[4]) #ACD通話
                                wb_sheet.cell(column=5, row=rows).value+=int(skill3[7]) #掛斷通話
                                holdsec+=int(skill3[20]) #等候時間
                                hold+=int(skill3[21]) #等候通話
                                TotalACD+=int(skill3[4])
                                talkingsec+=float(skill3[4])*float(skill3[5]) #ACD通話*平均ACD時間
                                acwsec+=float(skill3[4])*float(skill3[6]) #ACD通話*平均ACW時間
                                pickupsec+=float(skill3[4])*float(skill3[2]) #ACD通話*平均速度接聽
                                break
                    except: print(' => Missing Skill3', end="")
                        
                    #Skill4
                    try:
                        for skill4 in table_skill4_list:
                            if wb_sheet.cell(column=2, row=rows).value==Library.convertTimeDivisionFull(skill4[0], skill4[1]):
                                wb_sheet.cell(column=4, row=rows).value+=int(skill4[4]) #ACD通話
                                wb_sheet.cell(column=5, row=rows).value+=int(skill4[7]) #掛斷通話
                                holdsec+=int(skill4[20]) #等候時間
                                hold+=int(skill4[21]) #等候通話
                                TotalACD+=int(skill4[4])
                                talkingsec+=float(skill4[4])*float(skill4[5]) #ACD通話*平均ACD時間
                                acwsec+=float(skill4[4])*float(skill4[6]) #ACD通話*平均ACW時間
                                pickupsec+=float(skill4[4])*float(skill4[2]) #ACD通話*平均速度接聽
                                break
                    except: print(' => Missing Skill4', end="")

                    #VM
                    #try:
                    #    for vm in table_VM_list:
                    #        mydatetime=Library.convertDateTime(vm[0])
                    #        if mydatetime!=None: print(mydatetime)
                    #        if len(vm)>2 and wb_sheet.cell(column=2, row=rows).value==Library.convertTimeDivisionFullB(vm[0], vm[2]): #Csv前幾筆只有兩個欄位
                    #            wb_sheet.cell(column=6, row=rows).value+=int(vm[14]) #Voice Mail
                    #            break
                    #except: print(' => Missing VoiceMail', end="")
                    try:
                        for vm in table_VM_list:
                            mydatetime=Library.convertDateTime(vm[0])
                            if mydatetime!=None and Library.isSamePeriod(ReportDateStr + ' ' + wb_sheet.cell(column=2, row=rows).value, mydatetime):
                                wb_sheet.cell(column=6, row=rows).value+=1 #Voice Mail
                    except: print(' => Missing VoiceMail', end="")
                    
                    #Skill1--
                    Totaltalkingsec+=talkingsec
                    TotalHoldsec+=holdsec
                    TotalHold+=hold
                    TotalAcw+=acwsec
                    TotalPickupSec+=pickupsec
                    try: wb_sheet.cell(column=12, row=rows).value=(talkingsec/float(wb_sheet.cell(column=4, row=rows).value))/3600/24 #ACD通話*平均ACD時間
                    except: wb_sheet.cell(column=12, row=rows).value="--"
                    try: wb_sheet.cell(column=13, row=rows).value=(holdsec/hold)/3600/24 #ACD通話*平均ACD時間
                    except: wb_sheet.cell(column=13, row=rows).value="--"
                    try: wb_sheet.cell(column=14, row=rows).value=(holdsec/wb_sheet.cell(column=4, row=rows).value)/3600/24 #ACW通話*平均ACW時間
                    except: wb_sheet.cell(column=14, row=rows).value="--"
                    try: wb_sheet.cell(column=15, row=rows).value=(holdsec/wb_sheet.cell(column=4, row=rows).value)/3600/24 #ACD通話*平均速度接聽
                    except: wb_sheet.cell(column=15, row=rows).value="--"
            
                    print(' => Completed')
                except:
                    print(' => Failed')
            
            if TotalACD!=0: wb_sheet.cell(column=12, row=2).value=(Totaltalkingsec/TotalACD)/3600/24 #ACD通話/ACD
            if TotalHold!=0: wb_sheet.cell(column=13, row=2).value=(TotalHoldsec/TotalHold)/3600/24 #等候時間/等候通話
            if TotalACD!=0: wb_sheet.cell(column=14, row=2).value=(TotalAcw/TotalACD)/3600/24 #ACW通話/ACD
            if TotalACD!=0: wb_sheet.cell(column=15, row=2).value=(TotalPickupSec/TotalACD)/3600/24 #速度接聽/ACD

            print()
            print('Step 3: Generating Report')
            #Generate        
            wb_sheet.title = ReportDateSimple
            Performance_FilePathName = "Report\分時表%s.xlsx" % ReportDateSimple

            print('  -Creating Report to the %s'  % Performance_FilePathName, end="")

            wb.save(Performance_FilePathName)
            
            wb.close()
            print(' => Completed')  
        except Exception as e:
            print('  -Loading file: => Failed ')
            #print(e)
    except Exception as e:
        print('~Failed')
    #except:
    #   print("Error!! Close all excel files and try again.")

    finally:
        print()
        print('************ END ************')