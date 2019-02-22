# -*- coding: utf-8 -*-

from openpyxl import load_workbook
import csv
import datetime
import re
import os.path, os
import glob
import Library
from shutil import copyfile
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl import formatting, styles
from datetime import timedelta
from win32com.client import Dispatch
import shutil

def MainPerformance(rawdata_path, rawdata_date, withut, utrate, wordtimemarkup, debug):
    try:
        print('************ BEGIN ************')
        print()
        print('Step 1: Processing Template and Raw Data')
        try:
            
            #1.import template
            wb_file_name = 'Report\AG_Performance_Template.xlsx'
            wb_file = Library.getXlsxFile(wb_file_name, []) 
            print('  -Loading file: %s ' % wb_file, end="")
            #wb = load_workbook(filename = wb_file_name, data_only=True)
            wb = load_workbook(filename = wb_file) 
            for checksheet in wb.sheetnames:
                if checksheet!="Performance":
                    wb.remove(wb[checksheet])
            wb_sheet = wb["Performance"]
            ReportDate = rawdata_date if rawdata_date!=None else wb_sheet.cell(column=1, row=1).value
            wb_sheet.cell(column=1, row=1).value=rawdata_date
            #ReportDate = wb_sheet.cell(column=1, row=1).value
            ReportDateLast = ReportDate - datetime.timedelta(days=1)
            ReportDateSimpleArr = str(ReportDate).split(' ',1)[0].split('-',2)
            ReportDateSimple = ReportDateSimpleArr[1]+ReportDateSimpleArr[2]
            ReportDateStr = ReportDateSimpleArr[0]+"/"+ReportDateSimpleArr[1]+"/"+ReportDateSimpleArr[2]
            ReportDateStrNoZero = ReportDateSimpleArr[0]+"/"+str(int(ReportDateSimpleArr[1]))+"/"+str(int(ReportDateSimpleArr[2]))
            ReportDateLastSimpleArr = str(ReportDateLast).split(' ',1)[0].split('-',2)
            ReportDateLastSimple = ReportDateLastSimpleArr[1]+ReportDateLastSimpleArr[2]
            #wb_sheet = wb[wb.sheetnames[0]]
            print(' => Completed')
            
            #2.import data source
            print('  -Loading file:', end="")
            try:
                table_summary_file_name='客服人員群組總結'                
                table_summary_file, table_summary_list=Library.getCsvFile2(rawdata_path + '\**\*%s*.xls' % table_summary_file_name if rawdata_path!=None else None, 'Report\RAWDATA\**\*%s*.xls' % table_summary_file_name, [[1,1, ReportDateStrNoZero]])
                if table_summary_list!=None: print(' %s => Completed' % table_summary_file)
                else: print(' Finding "%s\" => Failed' % table_summary_file_name)
            except Exception as e:
                if debug: print(e)
                pass

            #2.import data source
            print('  -Loading file:', end="")
            try:
                table_loginout_file_name='客服人員登出登入'                
                table_loginout_file, table_loginout_list=Library.getCsvFile2(rawdata_path + '\**\*%s*.xls' % table_loginout_file_name if rawdata_path!=None else None, 'Report\RAWDATA\**\*%s*.xls' % table_loginout_file_name, [[1,1, ReportDateStrNoZero]])
                if table_loginout_list!=None: print(' %s => Completed' % table_loginout_file)
                else: print(' Finding "%s\" => Failed' % table_loginout_file_name)
            except Exception as e:
                pass

            
            #2.import data source
            print('  -Loading file:', end="")
            try:
                table_mail_file_name='MAIL'                
                table_mail_file, table_mail_sheet=Library.getXlsxFile2(rawdata_path + '\**\%s*.xlsx' % table_mail_file_name if rawdata_path!=None else None, 'Report\RAWDATA\**\*%s*.xlsx' % table_mail_file_name, [['Str', 2, 'Closed'], ['DateStr', 4, ReportDateStr]])
                if table_mail_file!=None: print(' %s => Completed' % table_mail_file)
                else: print(' Finding "%s\" => Failed' % table_mail_file_name)
            except Exception as e:
                pass

            #2.import data source
            print('  -Loading file:', end="")
            try:
                table_cts_file_name='CTS'                
                table_cts_file, table_cts_sheet=Library.getXlsxFile2(rawdata_path + '\**\*%s*.xlsx' % table_cts_file_name if rawdata_path!=None else None, 'Report\RAWDATA\**\*%s*.xlsx' % table_cts_file_name, [['DateStr', 3, ReportDateStr]])
                if table_cts_file!=None: print(' %s => Completed' % table_cts_file)
                else: print(' Finding "%s\" => Failed' % table_cts_file_name)
            except Exception as e:
                pass

            #2.import data source
            print('  -Loading file:', end="")
            try:
                table_survey_file_name=Library.getSurvey(ReportDate)
                #table_survey_file_name='Report\RAWDATA\滿意度調查-1227.xls'
                if table_survey_file_name==None: table_survey_file_name='滿意度調查'                
                table_survey_file, table_survey_list=Library.getCsvFile2(None, table_survey_file_name, [])
                if table_survey_list!=None: print(' %s => Completed' % table_survey_file)
                else: print(' Finding "%s\" => Failed' % table_survey_file_name)                
            except Exception as e:
                #print(e)
                pass            

            #2.import data source
            print('  -Loading file:', end="")
            try:
                table_working_file_name='%s年%s月班表' % (ReportDate.year, ReportDate.month)              
                table_working_file, table_working_sheet=Library.getXlsxFile2(rawdata_path + '\**\*%s*.xlsx' % table_working_file_name if rawdata_path!=None else None, 'Report\RAWDATA\**\*%s*.xlsx' % table_working_file_name, [])
                if table_working_file!=None: print(' %s => Completed' % table_working_file)
                else: print(' Finding "%s\" => Failed' % table_working_file_name)
            except Exception as e:
                pass

            #3.Processing Report
            print()
            print('Step 2: Processing report calculation')

            TotalACD = 0.0
            TotalACDCount = 0.0
            TotalACW = 0.0
            TotalACWCount = 0.0
            TotalACDTime = 0
            TotalACWTime = 0
            TotalRingTime = 0

            TableACDTime = 0
            TableACWTime = 0
            TableRingTime = 0
            TableLoginTime = 0 

            newTotalACD = 0.0
            newTotalACW = 0.0
            newTotalRing = 0.0
            newTotalLoginTime = 0.0
            newTotal14 = 0.0

            TotalLogin = 0

            if wb_sheet.max_row>1:
                wb_sheet.cell(column=4, row=3).value = 0

                #Process Summary Table 總計 Start
                mySummaryRow = Library.getRow(table_summary_list, '總計')
                if mySummaryRow!=None:
                    print('  -Loading LoginID = 總計', end="")
                    wb_sheet.cell(column=15, row=3).value = int(mySummaryRow[1]) #ACD通話
                    #wb_sheet.cell(column=13, row=3).value = datetime.timedelta(seconds=float(mySummaryRow[2]))   #Total ACDAVG
                    #wb_sheet.cell(column=14, row=3).value = datetime.timedelta(seconds=float(mySummaryRow[3]))   #Total ACWACG
                    TableACDTime=int(mySummaryRow[10]) #Total ACD
                    TableACWTime=int(mySummaryRow[11]) #Total ACW
                    TableRingTime=int(mySummaryRow[12]) #Total Ring
                    TableLoginTime=int(mySummaryRow[16]) #Total LoginTime
                    if TableLoginTime>0 and (TableACDTime+TableACWTime+TableRingTime)>0:
                        wb_sheet.cell(column=23, row=3).value = (TableACDTime+TableACWTime+TableRingTime)/TableLoginTime 

                    try: wb_sheet.cell(column=24, row=3).value = int(mySummaryRow[14])/int(mySummaryRow[16])   #文字服務/公務時間比例
                    except: wb_sheet.cell(column=24, row=3).value = 0

                if table_summary_file!=None and mySummaryRow!=None:
                    print(' => Completed')
                #Process Summary Table 總計 End

                for rows in range(1, wb_sheet.max_row+1):
                    try:
                        LoginID = wb_sheet.cell(column=4, row=rows).value
                        LoginName = wb_sheet.cell(column=2, row=rows).value           
                        
                        if table_summary_file!=None and table_summary_list!=None:                    
                            
                            mySummaryRow = Library.getRow(table_summary_list, LoginID)
                            if mySummaryRow==None: mySummaryRow = Library.getRow(table_summary_list, LoginName)
                                                
                            #Process Summary Table
                            if mySummaryRow!=None: print('  -Loading LoginID = %s'  % LoginID, end="")
                            if mySummaryRow!=None:                                

                                #print(mySummaryRow)
                                #Start ACD Process...
                                if int(mySummaryRow[1])>0: wb_sheet.cell(column=15, row=rows).value = int(mySummaryRow[1])   #ACD通話
                            
                            if mySummaryRow!=None and wb_sheet.cell(column=1, row=rows).value=='AG':   
                                TotalACD += float(mySummaryRow[2])
                                if float(mySummaryRow[2])>0: TotalACDCount+=1
                                wb_sheet.cell(column=13, row=rows).value = datetime.timedelta(seconds=float(mySummaryRow[2]))   #ACDAVG
                                wb_sheet.cell(column=13, row=3).value = datetime.timedelta(seconds=TotalACD/TotalACDCount)   #Total ACDAVG
                                #End ACD Process...
                                
                                #Start ACW Process...
                                TotalACW += float(mySummaryRow[3])
                                if float(mySummaryRow[3])>0: TotalACWCount+=1
                                wb_sheet.cell(column=14, row=rows).value = datetime.timedelta(seconds=float(mySummaryRow[3]))   #ACWAVG
                                wb_sheet.cell(column=14, row=3).value = datetime.timedelta(seconds=TotalACW/TotalACWCount)   #Total ACWACG
                                #End ACW Process...    

                                #Start Time Process...
                                try: 
                                    wb_sheet.cell(column=23, row=rows).value = (int(mySummaryRow[10])+int(mySummaryRow[11])+int(mySummaryRow[12]))/int(mySummaryRow[16])
                                    newTotalACD += int(mySummaryRow[10])
                                    newTotalACW += int(mySummaryRow[11])
                                    newTotalRing += int(mySummaryRow[12])
                                    newTotalLoginTime += int(mySummaryRow[16])
                                    wb_sheet.cell(column=23, row=3).value = (newTotalACD+newTotalACW+newTotalRing)/newTotalLoginTime

                                except: wb_sheet.cell(column=23, row=rows).value = 0
                                #End Time Process...    

                                #Start TotalLogin Process...
                                TotalLogin += int(mySummaryRow[16])
                                wb_sheet.cell(column=12, row=rows).value = datetime.timedelta(seconds=int(mySummaryRow[16]))   #TotalLogin
                                wb_sheet.cell(column=12, row=3).value = datetime.timedelta(seconds=TotalLogin)   #Total TotalLogin

                                try: 
                                    wb_sheet.cell(column=24, row=rows).value = int(mySummaryRow[14])/int(mySummaryRow[16])   #文字服務/公務時間比例
                                    newTotal14 += int(mySummaryRow[14]) #文字服務
                                    wb_sheet.cell(column=24, row=3).value = newTotal14/newTotalLoginTime   #文字服務/公務時間比例
                                except: wb_sheet.cell(column=24, row=rows).value = 0
                                
                                #End TotalLogin Process...

                                #Start ACH Process... Phaseout 20181230
                                #wb_sheet.cell(column=30, row=rows).value = datetime.timedelta(seconds=int(float(mySummaryRow[2])+float(mySummaryRow[3])))   #ACH
                                #End TotalLogin Process...

                                #Start 總處理工作時間 Process...2018/12/30
                                try:
                                    wb_sheet.cell(column=22, row=rows).value = int(mySummaryRow[16])/3600   #值班時間
                                except:
                                    wb_sheet.cell(column=22, row=rows).value = 0
                                #End TotalLogin Process...
                                                                                            
                        
                        #Process Loginout Table
                        if table_loginout_file!=None:
                            myLoginoutRow = Library.getRowMerge(table_loginout_list, LoginID)
                            if myLoginoutRow==None: myLoginoutRow = Library.getRowMerge(table_loginout_list, LoginName)
                            if myLoginoutRow!=None and wb_sheet.cell(column=1, row=rows).value=='AG':
                                #print(myLoginoutRow)
                                LoginSec = Library.getSec(myLoginoutRow[3])
                                LogoutSec = Library.getSec(myLoginoutRow[5])
                                wb_sheet.cell(column=10, row=rows).value = str(datetime.timedelta(seconds=LoginSec))[-5:]   #Login
                                wb_sheet.cell(column=11, row=rows).value = str(datetime.timedelta(seconds=LogoutSec))[-5:]   #Logout
                                try:
                                    #wb_sheet.cell(column=5, row=rows).value = int((LogoutSec-LoginSec)/60)
                                    wb_sheet.cell(column=5, row=rows).value = round(LoginSec/60, 0)
                                except Exception as e:
                                    wb_sheet.cell(column=5, row=rows).value = ""
                                    #print(e)
                            else:
                                myRole = wb_sheet.cell(column=1, row=rows).value
                                if myRole=='AG' or myRole=='SA':
                                    wb_sheet.cell(column=5, row=rows).value = "休"
                                    #wb_sheet.cell(column=22, row=rows).value = "休"
                                elif myRole=='管理職':
                                    wb_sheet.cell(column=5, row=rows).value = 9


                        #Process Mail Table
                        TotalPaperCounter = 0
                        if rows>3 and table_mail_file!=None:
                            myMailCounter = Library.getMailCount(table_mail_sheet, LoginName, ReportDateStr)                    
                            if myMailCounter!=None and myMailCounter>0:
                                wb_sheet.cell(column=19, row=rows).value = myMailCounter
                                TotalPaperCounter += myMailCounter   

                        #Process Cts Table
                        if rows>3 and table_cts_file!=None:
                            myFacebookCounter = Library.getCtsCount(table_cts_sheet, LoginName, ReportDateStr, 'Facebook')
                            if myFacebookCounter!=None and myFacebookCounter>0:
                                wb_sheet.cell(column=20, row=rows).value = myFacebookCounter
                                TotalPaperCounter += myFacebookCounter
                            
                            myTelCounter = Library.getCtsCount(table_cts_sheet, LoginName, ReportDateStr, '電話')
                            if myTelCounter!=None and myTelCounter>0:
                                wb_sheet.cell(column=16, row=rows).value = myTelCounter
                                TotalPaperCounter += myTelCounter

                            myOutboundCounter = Library.getCtsCount(table_cts_sheet, LoginName, ReportDateStr, 'Outbound')
                            if myOutboundCounter!=None and myOutboundCounter>0:
                                wb_sheet.cell(column=18, row=rows).value = myOutboundCounter 
                                TotalPaperCounter += myOutboundCounter

                        #Process Survey Table
                        if rows>3 and table_survey_list!=None:
                            TotalSurveyNo1, TotalSurveyNo2, TotalSurveyNo1Good, TotalSurveyNo2Good, TotalSurveyNo1Ans, TotalSurveyNo2Ans, TotalSurveyBothAns=Library.getSurveyCount(table_survey_list, LoginID)
                            try: wb_sheet.cell(column=26, row=rows).value=TotalSurveyNo1/int(wb_sheet.cell(column=15, row=rows).value)
                            except: pass
                            try: wb_sheet.cell(column=27, row=rows).value=TotalSurveyBothAns/int(wb_sheet.cell(column=15, row=rows).value)
                            except: pass
                            try: wb_sheet.cell(column=28, row=rows).value=TotalSurveyNo1Good/TotalSurveyNo1
                            except: pass
                            try: wb_sheet.cell(column=29, row=rows).value=TotalSurveyNo2Good/TotalSurveyNo2
                            except: pass

                        #Process 班表
                        if rows>3 and table_working_sheet!=None:
                            if wb_sheet.cell(column=5, row=rows).value == "休" and TotalPaperCounter>0:
                                wb_sheet.cell(column=5, row=rows).value = ""
                            if table_working_sheet!=None:
                                wb_sheet.cell(column=5, row=rows).value=Library.getWorkingHour(table_working_sheet, LoginName, ReportDate)
                         
                        #if str(wb_sheet.cell(column=5, row=rows).value).replace('.','',1).isdigit():
                        if Library.isNumber(wb_sheet.cell(column=5, row=rows).value)==1:
                            wb_sheet.cell(column=4, row=3).value += 1  
                        if table_summary_file!=None and mySummaryRow!=None:
                            print(' => Completed')
                    
                    except Exception as e:
                        print(' => Failed')
                        #print(e)    

                #Process 文字時間 Markup Start   
                try:  myUTRate=(newTotalACD+newTotalACW+newTotalRing+newTotal14)/newTotalLoginTime  
                except Exception as e: myUTRate=9999

                if withut and utrate!=0 and myUTRate<utrate:
                    print('  -Processing UT Rate check and markup, UTRate=%s, Markup=%s' % (utrate, wordtimemarkup), end="")
                    newTotal14=0 #Reset Total
                    for rows in range(1, wb_sheet.max_row+1):
                        try:
                            LoginID = wb_sheet.cell(column=4, row=rows).value
                            LoginName = wb_sheet.cell(column=2, row=rows).value                
                            
                            if table_summary_file!=None and table_summary_list!=None:                    
                                
                                mySummaryRow = Library.getRow(table_summary_list, LoginID)
                                              
                                if mySummaryRow!=None and wb_sheet.cell(column=1, row=rows).value=='AG':

                                    try: 
                                        texttime=int(float(mySummaryRow[14])*wordtimemarkup)
                                        if (int(mySummaryRow[10])+int(mySummaryRow[11])+int(mySummaryRow[12])+texttime)>int(mySummaryRow[16]): continue
                                        wb_sheet.cell(column=24, row=rows).value = texttime/int(mySummaryRow[16])   #文字服務/公務時間比例
                                        newTotal14 += texttime #文字服務
                                        wb_sheet.cell(column=24, row=3).value = newTotal14/newTotalLoginTime   #文字服務/公務時間比例
                                    except Exception as e:
                                        #print(e)  
                                        wb_sheet.cell(column=24, row=rows).value = 0
                                    
                                    #End TotalLogin Process...
                        
                        except Exception as e:
                            print(' => Failed')
                            #print(e)  
                    print(' => Completed')
                #Process 文字時間 Markup End

            #改數值格式後不用計算
            #if TotalLogin>0 and TotalACD>0:
            #    wb_sheet.cell(column=21, row=3).value = TotalACD/TotalLogin   # Total TotalACD/TotalLogin
            #if TotalLogin>0 and (TotalACDTime+TotalACWTime+TotalRingTime)>0:
            #    wb_sheet.cell(column=23, row=3).value = (TotalACDTime+TotalACWTime+TotalRingTime)/TotalLogin   # Total TotalACD/TotalLogin
            
            print()
            print('Step 3: Generating Report')
            #Generate        
            wb_sheet.title = ReportDateSimple
            Performance_FilePathName = "Report\OPPO_Agent_Performance%s.xlsx" % ReportDateSimple
            Performance_FilePathNameLast = "Report\OPPO_Agent_Performance%s.xlsx" % ReportDateLastSimple

            print('  -Creating Report to the %s'  % Performance_FilePathName, end="")

            gray_font = styles.Font(color='00A0A0A0')
            if not os.path.isfile(Performance_FilePathName) and os.path.isfile(Performance_FilePathNameLast) and ReportDateSimple[-2:]!='01':
                copyfile(Performance_FilePathNameLast, Performance_FilePathName)

            #if os.path.isfile(Performance_FilePathName):
            #    wb_Copy = load_workbook(filename = Performance_FilePathName)
            #    if ReportDateSimple in wb_Copy.sheetnames:
            #        wb_Copy_Sheet = wb_Copy[ReportDateSimple]    
            #    else:                    
            #        wb_Copy_Sheet = wb_Copy.copy_worksheet(wb_Copy[wb_Copy.sheetnames[len(wb_Copy.sheetnames)-1]])
            #        wb_Copy_Sheet.title = ReportDateSimple
            #    wb_Copy_Sheet = Library.copyWorksheet(wb_sheet, wb_Copy_Sheet)
                
            #    wb_Copy_Sheet.conditional_formatting.add('A1:AB100', CellIsRule(operator='equal', formula=['0'], stopIfTrue=True, font=gray_font))            
            #    wb_Copy.active = len(wb_Copy.sheetnames)-1
            #    wb_Copy.save(Performance_FilePathName)
            #    wb_Copy.close()

            #else:
            wb_sheet.conditional_formatting.add('A1:AB100', CellIsRule(operator='equal', formula=['0'], stopIfTrue=True, font=gray_font))
            wb.save(Performance_FilePathName)

            wb.close()
            Library.correctWorksheet(Performance_FilePathName)
            print(' => Completed')

        except Exception as e:
            print('  -Loading file: => Failed ')
            #print(e)

        try:
            print()
            print('Step 4: Generating Monthly Report')


            file_report_date=ReportDate
            file_report_firstdate=file_report_date.replace(day=1)
            file_report_currentdate=file_report_date

            Performance_FilePathName = "Report\OPPO_Agent_Performance%s_全月.xlsx" % ReportDateSimple
            Performance_FilePathName_Full=os.path.abspath(Performance_FilePathName)
            if os.path.isfile(wb_file_name):
                shutil.copy2(wb_file_name, Performance_FilePathName)

            xl = Dispatch("Excel.Application")
            xl.Visible = False
            xl.DisplayAlerts = False
            while (file_report_currentdate>=file_report_firstdate):
                MonthlyReportDateSimpleArr = str(file_report_currentdate).split(' ',1)[0].split('-',2)
                MonthlyReportDateSimple = MonthlyReportDateSimpleArr[1]+MonthlyReportDateSimpleArr[2]
                source=Library.getFilePath(rawdata_path+'\**\OPPO_Agent_Performance%s.xlsx' % MonthlyReportDateSimple, 'Report\RAWDATA\**\OPPO_Agent_Performance%s.xlsx'  % MonthlyReportDateSimple)
                print('  -Loading Date=%s, File=%s' % (MonthlyReportDateSimple,source), end="")
                if source!=None:
                    wb1 = xl.Workbooks.Open(Filename=source)
                    wb2 = xl.Workbooks.Open(Filename=Performance_FilePathName_Full)
                    ws1 = wb1.Worksheets(1)
                    ws1.Copy(Before=wb2.Worksheets(1))
                    wb2.Close(SaveChanges=True)
                    wb2=None
                    wb1.Close(SaveChanges=False)
                    wb1=None

                file_report_currentdate=file_report_currentdate - timedelta(days=1)
                print(' => Completed')

            wb2 = xl.Workbooks.Open(Filename=Performance_FilePathName_Full)
            try: 
                wb2_sheet = wb2.Sheets("SOP")
                wb2_sheet.Delete()
            except: pass
            try:
                wb2_sheet = wb2.Sheets("Performance")
                wb2_sheet.Delete()
            except: pass
            try: 
                wb2_sheet = wb2.Sheets("分時表")
                wb2_sheet.Delete()
            except: pass
            try: 
                wb2_sheet = wb2.Sheets("問題解決率")
                wb2_sheet.Delete()
            except: pass

            wb2.Close(SaveChanges=True)
            wb2=None

            xl.Quit()
            xl = None

        except Exception as e:
            print(' => Failed ')
            #print(e)

    #except Exception as e:
    #    print(e)
    except Exception as e:
       print("Error!! Close all excel files and try again.")

    finally:
        print()
        print('************ END ************')