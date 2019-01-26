# -*- coding: utf-8 -*-

from openpyxl import load_workbook
import datetime
import os.path
import glob
import Library
from shutil import copyfile

def MainInbound(rawdata_path):
    try:
        print('************ BEGIN ************')
        print()
        print('Step 1: Processing Template and Raw Data')
        try:
            
            #1.import template
            wb_file_name = 'Report\AG_Performance_Template.xlsx'
            wb_file = Library.getXlsxFile(wb_file_name, [])
            print('  -Loading file: %s ' % wb_file, end="")
            wb = load_workbook(filename = wb_file)
            wb_sheet = wb["Performance"]
            
            for checksheet in wb.sheetnames:
                if checksheet!="問題解決率":
                    wb.remove(wb[checksheet])
            wb_sheet = wb["問題解決率"]
            print(' => Completed')

            #2.import data source
            #Skill1
            print('  -Loading file:', end="")
            try:
                table_inbound_file_name=rawdata_path.replace('.xlsx', '')          
                table_inbound_file, table_inbound_sheet=Library.getXlsxFile2(rawdata_path + '\**\*%s*.xlsx' % table_inbound_file_name if rawdata_path!=None else None, 'Report\RAWDATA\**\*%s*.xlsx' % table_inbound_file_name, [])
                if table_inbound_file!=None: print(' %s => Completed' % table_inbound_file)
                else: print(' Finding "%s\" => Failed' % table_inbound_file_name)
            except Exception as e:
                #print(e)
                pass

            #3.Processing Report
            print()
            print('Step 2: Processing report calculation')

            #wb_sheet.cell(column=1, row=2).value=ReportDateStr
            email_count=0
            call_count=0
            facebook_count=0
            email_repeat_count=0
            call_repeat_count=0
            facebook_repeat_count=0
            print('  -Getting base date:', end="")
            firstDate=Library.getFirstDate(table_inbound_sheet)
            print(' %s' %firstDate)
            
            wb_sheet_RepeatRawData=wb.create_sheet('RepeatRawData')
            #wb_sheet_RepeatRawData.append(wb_sheet_InboundListRawData.iter_rows(1))
            wb_sheet_RepeatRawData=Library.copyWorksheetRow(table_inbound_sheet, 1, wb_sheet_RepeatRawData)

            #ws.delete_rows(index, 1)
            #wb_sheet_InboundListRawData.column_dimensions['E'].number_format='hh:mm:ss'
            for rows in range(2, table_inbound_sheet.max_row+1): 
                try:
                    #print('  -Progressing of data is %s/%s [%s], ID=%s, 進線管道=%s, 客戶姓名=%s' % (rows, table_inbound_sheet.max_row, "{0:.2%}".format((rows)/table_inbound_sheet.max_row), table_inbound_sheet.cell(column=1, row=rows).value, table_inbound_sheet.cell(column=2, row=rows).value, table_inbound_sheet.cell(column=5, row=rows).value), end="")
                    #wb_sheet_InboundListRawData.cell(row=rows, column=5).number_format='hh:mm:ss'
                    if table_inbound_sheet.cell(column=1, row=rows).value==None: continue
                    print('#', end="")
                    if rows%50==0 or rows==table_inbound_sheet.max_row: 
                        print(" %s/%s [%s] => Completed" % (rows, table_inbound_sheet.max_row, "{0:.2%}".format((rows)/table_inbound_sheet.max_row)))
                    
                    if Library.convertDate(table_inbound_sheet.cell(column=3, row=rows).value)<firstDate: continue
                    if table_inbound_sheet.cell(column=2, row=rows).value=='email': email_count+=1
                    elif table_inbound_sheet.cell(column=2, row=rows).value=='電話': call_count+=1
                    elif table_inbound_sheet.cell(column=2, row=rows).value=='Facebook': facebook_count+=1
                    for searchrows in range(2, table_inbound_sheet.max_row+1):
                        try:
                            if table_inbound_sheet.cell(column=1, row=searchrows).value==None: continue   
                            if ( 
                            table_inbound_sheet.cell(column=19, row=rows).value!='錯誤來電' and
                            table_inbound_sheet.cell(column=1, row=rows).value!=table_inbound_sheet.cell(column=1, row=searchrows).value and #非同一筆
                            table_inbound_sheet.cell(column=2, row=rows).value==table_inbound_sheet.cell(column=2, row=searchrows).value and #進線管道
                            table_inbound_sheet.cell(column=5, row=rows).value==table_inbound_sheet.cell(column=5, row=searchrows).value and #客戶姓名
                            table_inbound_sheet.cell(column=7, row=rows).value==table_inbound_sheet.cell(column=7, row=searchrows).value and #聯絡電話
                            table_inbound_sheet.cell(column=13, row=rows).value==table_inbound_sheet.cell(column=13, row=searchrows).value and #諮詢分類1
                            table_inbound_sheet.cell(column=14, row=rows).value==table_inbound_sheet.cell(column=14, row=searchrows).value and #諮詢分類2
                            table_inbound_sheet.cell(column=15, row=rows).value==table_inbound_sheet.cell(column=15, row=searchrows).value and #諮詢分類3
                            table_inbound_sheet.cell(column=16, row=rows).value==table_inbound_sheet.cell(column=16, row=searchrows).value and #諮詢分類4
                            0<(table_inbound_sheet.cell(column=3, row=rows).value-table_inbound_sheet.cell(column=3, row=searchrows).value).total_seconds()<=60*60*48 #超過48小時
                            ):
                                if table_inbound_sheet.cell(column=2, row=rows).value=='email': email_repeat_count+=1
                                elif table_inbound_sheet.cell(column=2, row=rows).value=='電話': call_repeat_count+=1
                                elif table_inbound_sheet.cell(column=2, row=rows).value=='Facebook': facebook_repeat_count+=1
                                #print((table_inbound_sheet.cell(column=3, row=rows).value-table_inbound_sheet.cell(column=3, row=searchrows).value).total_seconds())
                                if table_inbound_sheet.cell(column=2, row=rows).value in ('email','電話','Facebook'):                            
                                    wb_sheet_RepeatRawData=Library.copyWorksheetRow(table_inbound_sheet, searchrows, wb_sheet_RepeatRawData)
                                    wb_sheet_RepeatRawData=Library.copyWorksheetRow(table_inbound_sheet, rows, wb_sheet_RepeatRawData)
                                break
                                
                        except Exception as e:
                            print(' => Failed')
                            pass
                    
                    #print(' => Completed') 

                except Exception as e:
                            print(e)
                            pass

            #for row in range(1, wb_sheet_RepeatRawData.max_row):
            #    wb_sheet_RepeatRawData["E{}".format(row)].number_format = 'hh:mm:ss'
            
            wb_sheet.cell(column=2, row=3).value=call_count
            wb_sheet.cell(column=2, row=4).value=call_repeat_count
            wb_sheet.cell(column=3, row=3).value=email_count
            wb_sheet.cell(column=3, row=4).value=email_repeat_count
            wb_sheet.cell(column=4, row=3).value=facebook_count
            wb_sheet.cell(column=4, row=4).value=facebook_repeat_count
            #print(table_inbound_sheet)           

            #Generate
            print()
            print('Step 3: Generating Report')

            #Generate RawData

            #Copy RawData to target
            print('  -Coping RawData to the Sheet', end="")
            wb_sheet_InboundListRawData=wb.create_sheet('InboundListRawData')
            wb_sheet_InboundListRawData=Library.copyWorksheet(table_inbound_sheet, wb_sheet_InboundListRawData)
            print(' => Completed') 

            wb.active = 0
            #wb_sheet_RepeatRawData.conditional_formatting.add('E:E', CellIsRule(operator='equal', formula=['0'], stopIfTrue=True, font=gray_font))
                 
            Inboundlist_FilePathName = "Report\問題解決率-%s.xlsx" % rawdata_path.replace('.xlsx', '')

            print('  -Creating Report to the %s'  % Inboundlist_FilePathName, end="")

            wb.save(Inboundlist_FilePathName)
            
            wb.close()
            print(' => Completed') 
        except Exception as e:
            print('  -Loading file: => Failed ')
            print(e)
    except Exception as e:
        print('~Failed')
    #except:
    #   print("Error!! Close all excel files and try again.")

    finally:
        print()
        print('************ END ************')