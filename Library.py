# -*- coding: utf-8 -*-
from openpyxl import load_workbook
import csv
import datetime
import re
import os.path
import glob
from shutil import copyfile
import configparser
import urllib3
from openpyxl.styles import Font

def getRow(table, loginID):
    for i in table:
        if formatLoginID(i[0])==formatLoginID(loginID):
        #if i[0][-7:]==str(myloginID):
            return i
    return None

def formatLoginID(loginID):
    
    try:
        myloginID = str(loginID)
        if len(myloginID)==7:
            myloginID = 'A'+myloginID
        return myloginID
    except Exception as e:
        #print(e)
        return loginID

def getRowMerge(table, loginID):
    myloginID = str(loginID)
    if len(myloginID)==7:
        myloginID = 'A'+myloginID
    myTable=None
    myStart=None
    first=0
    try:
        for i in table:        
            if i[0]==myloginID:
                if first==0: myStart=i[3]
                first=1
                myTable=i
        if myTable!=None and myStart!=None:
            myTable[3]=myStart
        return myTable
    except Exception as e:
        #print(e)
        return None

def getSetting(section, field, checkpath, convertdate, checkfloat):
    try:
        config = configparser.ConfigParser()
        config.read('Settings.ini', encoding='ANSI')
        result=config.get(section, field)
        if result=="": result=None
        else: 
            if checkpath and not os.path.isdir(result): 
                trylist=['Z','Y','X','W','V','U','T','S','R','Q','P','O','N','M','L','K','J','I','H','G','F','E']
                for i in trylist:
                    result=i+result[1:]
                    if os.path.isdir(result): return result
                    else: result=None
                #if checkpath and not os.path.isdir(result): result=None
            if convertdate: result=datetime.datetime.strptime(result, "%Y/%m/%d").date()
            if checkfloat: result=float(result)
        return result
    except Exception as e:
        #print(e)        
        return None

def formatDateToStr(importDate):
    return str(importDate).split(' ',1)[0].replace("-", "/")

def convertDate(mydate):
    try:
        mydatestr=str(mydate).split(' ',1)[0].replace('-', '/')
        result=datetime.datetime.strptime(mydatestr, "%Y/%m/%d").date()
        return result
    except Exception as e:
        #print(e)
        return None

def convertDateTime(mydatetime):
    try:
        mydatetimestr=str(mydatetime).replace('/', '-')
        result=datetime.datetime.strptime(mydatetimestr, "%Y-%m-%d %H:%M:%S.%f")
        return result
    except Exception as e:
        #print(e)
        return None

def isSamePeriod(source, target):
    try:
        source_startdatetime=str(source).split('-')[0]
        source_startdatetime=datetime.datetime.strptime(source_startdatetime, "%Y/%m/%d %H:%M")
        #if source_startdatetime.minute==30: source_startdatetime=source_startdatetime
        source_enddatetime=str(source).split(' ')[0]+ ' ' + str(source).split('-')[1]
        source_enddatetime=datetime.datetime.strptime(source_enddatetime, "%Y/%m/%d %H:%M")
        if target>=source_startdatetime and target<=source_enddatetime: return 1
        return 0
    except Exception as e:
        #print(e)
        return 0

def getMailCount(table, loginName, reportDate):
    try:
        myCount = 0
        #reportDate.replace(hour=0, minute=0, second=0, microsecond=0)
        if table.max_row>1 and loginName!=None:
            for rows in range(2, table.max_row+1):            
                myLoginName = table.cell(column=5, row=rows).value
                if myLoginName!=None:
                    myLoginName = re.sub(r'[\x00-\x7f]',r' ', myLoginName).strip()
                    reportDate = convertDate(reportDate)
                    myMailDate = convertDate(table.cell(column=4, row=rows).value)
                    #myMailDate = myMailDate.replace(hour=0, minute=0, second=0, microsecond=0)
                    if table.cell(column=2, row=rows).value=='Closed' and loginName==myLoginName and reportDate==myMailDate: 
                        myCount += 1
        else:
            return None   
        
        return myCount if myCount>0 else None        
    except:
        return None

def getCtsCount(table, loginName, reportDate, serviceWay):
    try:
        myCount = 0
        #myReportDate = formatDateToStr(reportDate)
        if table.max_row>1 and loginName!=None and serviceWay!=None and reportDate!=None:
            for rows in range(2, table.max_row+1):            
                myLoginName = str(table.cell(column=20, row=rows).value).replace('　', '').strip()
                #print('Cts Name=%s' % myLoginName)
                if myLoginName!=None:
                    LoginName = re.sub(r'[\x00-\x7f]',r' ', loginName).strip()
                    reportDate = convertDate(reportDate)
                    myMailDate = convertDate(table.cell(column=3, row=rows).value)
                    #print('LoginName=%s, myLoginName=%s, reportDate=%s, myMailDate=%s, serviceWay1=%s, serviceWay2=%s' % (LoginName, myLoginName, myReportDate, myMailDate, table.cell(column=2, row=rows).value, serviceWay))
                    if table.cell(column=2, row=rows).value==serviceWay and LoginName==myLoginName and reportDate==myMailDate: 
                        myCount += 1
        else:
            return None
            
        return myCount if myCount>0 else None
    except Exception as e:
        print(' => CTS Data Abnormal', end="")
        #print(e)
        return None

def getWorkingHour(table, loginName, reportDate):
    try:
        myCol=0
        myRow=0
        myValue=None
        for rows in range(1, table.max_row+1):
            for cols in range(1, table.max_column):
                if isinstance(table.cell(column=cols, row=rows).value, datetime.datetime) and table.cell(column=cols, row=rows).value.date()==reportDate:
                    #print("rows=%s, cols=%s"%(rows,cols))
                    myCol=cols                
                if loginName=='崔鳳婷' and str(table.cell(column=cols, row=rows).value)=='崔鳯婷': myRow=rows #字打錯了
                if str(table.cell(column=cols, row=rows).value)==loginName: myRow=rows
                if myCol!=0 and myRow!=0: break
        if myCol!=0 and myRow!=0:
            myValue=str(table.cell(column=myCol, row=myRow).value)
            if isNumber(myValue):
                myValue=int(myValue[:2])
        #print("loginName=%s, myValue=%s, myCol=%s, myRow=%s"%(loginName,myValue,myCol,myRow))
        return myValue
    except Exception as e:
        print(e)
        return None

def getSec(iMinSec):
    try:
        minsec = iMinSec[:-2].split(':',1)
        return int(minsec[1])+(int(minsec[0])*60)+(12*60 if iMinSec[-2:]=='下午' else 0)
    except:
        print(' => Login/out Time Abnormal', end="")
        return 0

def convertTimeDivision(segment):
    try:
        minstr = str(segment)[-2:]
        hourstr = str(segment)[:-2]
        if hourstr == '': hourstr = '00'
        if minstr == '60':
            hourstr = str(int(hourstr)+1)
            minstr = '00'
        return hourstr.zfill(2) + ":" + minstr.zfill(2)
    except:
        return '00:00-00:00'

def convertTimeDivisionFull(start, end):
    return convertTimeDivision(start) + "-" + convertTimeDivision(end)

def convertTimeDivisionFullB(start, end):
    try:
        starthour = int(start.split(':',1)[0])
        startmin  = int(start.split(':',1)[1])
        endhour   = int(end.split(':',1)[0])
        endmin    = int(end.split(':',1)[1][:-2])
        if end.split(':',1)[1][-2:]=='上午' and endhour==12: endhour = starthour = 0
        if end.split(':',1)[1][-2:]=='下午' and starthour<12 and not (starthour==11 and startmin==30): starthour += 12
        if end.split(':',1)[1][-2:]=='下午' and endhour<12: endhour += 12
        if end=='12:00上午': 
            starthour = 11
            endhour = 24
        return str(starthour).zfill(2) + ":" + str(startmin).zfill(2) + "-" + str(endhour).zfill(2) + ":" + str(endmin).zfill(2)
    except:
        return '00:00-00:00'

def copyWorksheet(source, target):
    for rows in range(1, source.max_row+100):
        for cols in range(1, source.max_column+1):
            target.cell(column=cols, row=rows).value = source.cell(column=cols, row=rows).value            
            #if target.cell(column=cols, row=rows).value==0:
            #    target.cell(column=cols, row=rows).style.font.color.index = Color.Gray
    return target

def copyWorksheetRow(source, index, target):
    newRow=target.max_row+1 if index!=1 else 1
    for cols in range(1, source.max_column+1):
        target.cell(column=cols, row=newRow).value = source.cell(column=cols, row=index).value
        #target.cell(column=cols, row=newRow).font = Font(underline="single")
    
    #target.row_dimensions[newRow].font = Font(underline="single")
    return target

def isNumber(myValue):
    try:
        test=str(myValue)
        if test.replace('.','').isdigit()==1: return 1
        else: return 0
    except: return 0

def correctWorksheet(filepath):
    try:
        #print(filepath)
        if os.path.isfile(filepath):
            wb = load_workbook(filename = filepath)
            for mySheet in wb:
                #print(mySheet)
                #print(mySheet.cell(column=13, row=5).value)
                #print(mySheet.cell(column=12, row=6).value)
                for rows in range(1, mySheet.max_row+1):
                    for cols in range(1, mySheet.max_column+1):
                        if str(mySheet.cell(column=cols, row=rows).value)[:5]=="1899-":  
                            mySheet.cell(column=cols, row=rows).value=0
                        #elif str(mySheet.cell(column=cols, row=rows).value)=="1899-12-31 00:00:00" or str(mySheet.cell(column=cols, row=rows).value)=="1899-12-21 00:00:00": mySheet.cell(column=cols, row=rows).value=1
            wb.save(filepath)
            wb.close()
    except Exception as e:
        print(' => File Lock')

def getSurvey(myDate):
    try:
        http = urllib3.PoolManager()
        headers = {'User-Agent': 'Mozilla/5.0'}
        myYear=str(myDate.year)
        myMonth=str(myDate.month).zfill(2)
        myDay=str(myDate.day).zfill(2)
        datesVal1=myYear+"/"+myMonth+"/"+myDay
        datesVal2=myYear+myMonth+myDay
        data = {'form1': '?do=QA', 'subject': 'All', 'level': 'All', 'submit': '產生報表', 's_auditw_sY': myYear, 's_auditw_sM': myMonth, 's_auditw_sD': myDay, 'tH': '0', 'tM': '0', 'tS': '0', 's_auditw_eY': myYear, 's_auditw_eM': myMonth, 's_auditw_eD': myDay, 'gH': '23', 'gM': '59', 'gS': '59','datesVal1': datesVal1, 'datesVal2': datesVal2, 'dateeVal1': datesVal1, 'dateeVal2': datesVal2}
        webcontent = http.request('POST', 'http://10.48.24.37/webreport/satis.php?do=QA',headers=headers , fields=data, timeout=30)
        myData=webcontent.data.decode('cp950')
        #oData= Selector(text=myData)
        #filename=oData.xpath('//a//@href').extract()[0][1:]
        filename=myData.split('href=')[2].split('.')[1]+'.xls'
        filenamefull='http://10.48.24.37/webreport'+filename
        targetfile='滿意度調查-'+ myMonth + myDay +'.xls'
        targetfilefull='Report\RAWDATA\\'+targetfile
        response = http.request('GET', filenamefull)
        with open(targetfilefull, 'wb') as f:
            f.write(response.data)

        response.release_conn()
        webcontent.release_conn()

        return targetfilefull

    except Exception as e:
        #print(e)
        return None

def getVoiceMail(myDate):
    try:
        http = urllib3.PoolManager()
        headers = {'User-Agent': 'Mozilla/5.0'}
        myYear=str(myDate.year)
        myMonth=str(myDate.month).zfill(2)
        myDay=str(myDate.day).zfill(2)
        datesVal1=myYear+"/"+myMonth+"/"+myDay
        datesVal2=myYear+myMonth+myDay
        data = {'form1': '?do=QA', 'subject': 'All', 'level': 'All', 'submit': '產生報表', 's_auditw_sY': myYear, 's_auditw_sM': myMonth, 's_auditw_sD': myDay, 'tH': '0', 'tM': '0', 'tS': '0', 's_auditw_eY': myYear, 's_auditw_eM': myMonth, 's_auditw_eD': myDay, 'gH': '23', 'gM': '59', 'gS': '59','datesVal1': datesVal1, 'datesVal2': datesVal2, 'dateeVal1': datesVal1, 'dateeVal2': datesVal2}
        webcontent = http.request('POST', 'http://10.48.24.37/webreport/leavePhoneNo.php?do=QA',headers=headers , fields=data, timeout=30)
        myData=webcontent.data.decode('cp950')
        #oData= Selector(text=myData)
        #filename=oData.xpath('//a//@href').extract()[0][1:]
        filename=myData.split('href=')[3].split('.')[1]+'.xls'
        filenamefull='http://10.48.24.37/webreport'+filename
        targetfile='留電報表-'+ myMonth + myDay +'.xls'
        targetfilefull='Report\RAWDATA\\'+targetfile
        response = http.request('GET', filenamefull)
        with open(targetfilefull, 'wb') as f:
            f.write(response.data)

        response.release_conn()
        webcontent.release_conn()

        return targetfilefull

    except Exception as e:
        #print(e)
        return None
        

def xxxxgetCsvFile(file, conditions):
    for filename in glob.iglob(file, recursive=True):
        try:
            myTable = csv.reader(open(filename, 'r',encoding='ANSI'), delimiter='\t')
            myTable_list = list(myTable)
            if len(conditions)==0: return filename
            myFlag=0
            for conditon in conditions:
                if myTable_list[conditon[1]][conditon[0]]==conditon[2]: myFlag=1
                else: 
                    myFlag=0
                    break
            if myFlag: return filename
        except Exception as e:
            print(' => Data Exception')
            #print(e)
    return None

def getCsvFile2(file, filesecond, conditions):
    for times in range(0, 2):
        try:
            if times==0: myFile=file 
            else: myFile=filesecond
            if myFile==None: continue
            
            for filename in glob.iglob(myFile, recursive=True):       
                try:         
                    myTable = csv.reader(open(filename, 'r',encoding='ANSI'), delimiter='\t')
                    myTable_list = list(myTable) if myTable!=None else None
                    if len(conditions)==0: return filename, myTable_list
                    myFlag=0
                    for conditon in conditions:
                        if myTable_list[conditon[1]][conditon[0]]==conditon[2]: myFlag=1
                        else: 
                            myFlag=0
                            break
                    if myFlag: return filename, myTable_list
                except Exception as e:
                    pass
        
        except Exception as e:
            if times==0: myTable_list=None
            else: return None, None
            #print(e)
    return None, None

def getFilePath(file, filesecond):
    for times in range(0, 2):
        try:
            if times==0: myFile=file 
            else: myFile=filesecond
            if myFile==None: continue
            
            for filename in glob.iglob(myFile, recursive=True):             
                return filename
        
        except Exception as e:
            return None
            #print(e)
    return None

def getSurveyCount(surveyList, loginID):
    TotalSurveyNo1=0
    TotalSurveyNo1Ans=0
    TotalSurveyNo2=0
    TotalSurveyNo2Ans=0
    TotalSurveyNo1Good=0
    TotalSurveyNo2Good=0
    TotalSurveyBothAns=0
    lastrow=None
    for row in surveyList:
        if len(row)==8 and row[2]==str(loginID):
            if row[6]=='1': 
                TotalSurveyNo1+=1
                if row[5]=='非常滿意' or row[5]=='滿意':
                    TotalSurveyNo1Good+=1
                if row[5]!='': TotalSurveyNo1Ans+=1 #有做答筆數
            if row[6]=='2':
                TotalSurveyNo2+=1
                if row[5]=='非常不滿意' or row[5]=='非常滿意': 
                    TotalSurveyNo2Good+=1
                if row[5]!='': TotalSurveyNo2Ans+=1 #有做答筆數

                if lastrow!=None and row[5]!='' and lastrow[6]=='1' and lastrow[5]!='' and row[2]==lastrow[2] and row[4]==lastrow[4]:
                    TotalSurveyBothAns+=1 #兩題都有做答筆數
            lastrow=row
    
    #if str(loginID)=='2070608':
    #    print('TotalSurveyNo1=%s, TotalSurveyNo2Ans=%s' % (TotalSurveyNo1, TotalSurveyNo2Ans))
            
    return TotalSurveyNo1, TotalSurveyNo2, TotalSurveyNo1Good, TotalSurveyNo2Good, TotalSurveyNo1Ans, TotalSurveyNo2Ans, TotalSurveyBothAns


def getXlsxFile2(file, filesecond, conditions):
    for times in range(0, 2):
        try:
            if times==0: myFile=file
            else: myFile=filesecond
            if myFile==None: continue
            for filename in glob.iglob(myFile, recursive=True):
                try:
                    table = load_workbook(filename = filename) 
                    table_sheet = table[table.sheetnames[0]]                
                    if len(conditions)==0: return filename, table_sheet
                    for rows in range(2, table_sheet.max_row+1):                   
                        myFlag = 0                
                        for myCondition in conditions:
                            if myCondition[0] == 'DateStr' and convertDate(myCondition[2]) == convertDate(table_sheet.cell(column=myCondition[1], row=rows).value):
                                myFlag = 1
                            elif myCondition[0] == 'Str' and myCondition[2] == table_sheet.cell(column=myCondition[1], row=rows).value:
                                myFlag = 1
                            else:
                                myFlag = 0
                        if myFlag:
                            table.close()
                            return filename, table_sheet                
                    table.close()
                except Exception as e:
                    pass
        except Exception as e:
            #print(e)
            #print(e.code)
            #if e.code==13: print('a')
            if times==0: table_sheet=None
            else: return None, None            
    return None, None

def getXlsxFile(file, conditions):
    for filename in glob.iglob(file, recursive=True):
        try:
            #print(filename)
            table = load_workbook(filename = filename) 
            table_sheet = table[table.sheetnames[0]]
            if len(conditions)==0: return filename
            for rows in range(2, table_sheet.max_row+1):
                myFlag = 0                
                for myCondition in conditions:
                    if myCondition[0] == 'DateStr' and myCondition[2] == str(table_sheet.cell(column=myCondition[1], row=rows).value).split(' ',1)[0].replace("-", "/"):
                        myFlag = 1
                    elif myCondition[0] == 'Str' and myCondition[2] == table_sheet.cell(column=myCondition[1], row=rows).value:
                        myFlag = 1
                    else:
                        myFlag = 0
                if myFlag == 1:
                    table.close()
                    return filename                
            table.close()
        except Exception as e:
            table.close()
            #print(' => Data Exception')
            #print(e)
    return None

def getFirstDate(worksheet):
    try:
        myDate=[]
        for rows in range(2, worksheet.max_row+1):
            tempDate=convertDate(worksheet.cell(column=3, row=rows).value)
            if  tempDate not in myDate:
                myDate.append(tempDate)
        for rows in range(2, worksheet.max_row+1):
            result=convertDate(worksheet.cell(column=3, row=rows).value)
            if (len(myDate)>=20 and result.day==1) or (len(myDate)<20 and result.weekday()==0):                
                return result    
                    
    except Exception as e:
        #print(e)
        return None


def xxxxgetCsvFileName(file, myDate):
    for filename in glob.iglob('Report\RAWDATA\**\%s' % file, recursive=True):
        ##print(' %s' % filename, end="")
        try:
            myTable = csv.reader(open(filename, 'r'), delimiter='\t')
            myTable_list = list(myTable)
            if myTable_list[1][1]==myDate:
                ##print(' => Completed')
                return filename
        except:
            print(' => Data Exception')
    return None

def xxxxgetMailFileName(file, myDate):  
    for filename in glob.iglob('Report\RAWDATA\**\%s' % file, recursive=True):
        #print(' %s' % filename, end="")
        try:
            #print(filename)
            table = load_workbook(filename = filename) 
            table_sheet = table[table.sheetnames[0]]       
            for rows in range(2, table_sheet.max_row+1):  
                myFileDate = formatDateToStr(table_sheet.cell(column=4, row=rows).value)
                if table_sheet.cell(column=2, row=rows).value=='Closed' and myDate==myFileDate:
                    #print(' => Completed')
                    return filename
        except:
            print(' => Data Exception')
    return None

def xxxxgetCtsFileName(file, myDate):
    for filename in glob.iglob('Report\RAWDATA\**\%s' % file, recursive=True):
        #print(' %s' % filename, end="")
        try:
            table = load_workbook(filename = filename) 
            table_sheet = table[table.sheetnames[0]]
            for rows in range(2, table_sheet.max_row+1):  
                myFileDate = formatDateToStrtable_sheet.cell(column=3, row=rows).value
                if myDate==myFileDate:                    
                    return filename
        except:
            print(' => Data Exception')
        #except Exception as e:
        #    print(e)
    return None